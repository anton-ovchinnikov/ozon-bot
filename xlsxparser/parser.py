import openpyxl

CLUSTERS = {
    "ГРИВНО_РФЦ": "Москва-Запад",
    "ПЕТРОВСКОЕ_РФЦ": "Москва-Запад",
    "ХОРУГВИНО_РФЦ": "Москва-Запад",
    "НОГИНСК_РФЦ": "Москва-Восток",
    "ПУШКИНО_1_РФЦ": "Москва-Восток",
    "ПУШКИНО_2_РФЦ": "Москва-Восток",
    "СОФЬИНО_РФЦ": "Центр",
    "ТВЕРЬ_РФЦ": "Центр",
    "ТВЕРЬ_ХАБ": "Центр",
    "САНКТ_ПЕТЕРБУРГ_РФЦ": "Санкт-Петербург и СЗО",
    "СПБ_БУГРЫ_РФЦ": "Санкт-Петербург и СЗО",
    "СПБ_ШУШАРЫ_РФЦ": "Санкт-Петербург и СЗО",
    "КАЗАНЬ_РФЦ": "Поволжье",
    "НИЖНИЙ_НОВГОРОД_РФЦ": "Поволжье",
    "САМАРА_РФЦ": "Поволжье",
    "ВОРОНЕЖ_МРФЦ": "Дон",
    "ВОРОНЕЖ_2_РФЦ": "Дон",
    "ВОЛГОГРАД_МРФЦ": "Дон",
    "РОСТОВ_НА_ДОНУ_РФЦ": "Дон",
    "АДЫГЕЙСК_РФЦ": "Юг",
    "НОВОРОССИЙСК_МРФЦ": "Юг",
    "ЕКАТЕРИНБУРГ_РФЦ": "Урал",
    "КРАСНОЯРСК_МРФЦ": "Сибирь",
    "НОВОСИБИРСК_РФЦ": "Сибирь",
    "ХАБАРОВСК_РФЦ": "Дальний Восток",
    "ХАБАРОВСК_2_РФЦ": "Дальний Восток",
    "КАЛИНИНГРАД_МРФЦ": "Калининград",
    "АСТАНА_РФЦ": "Казахстан",
    "АЛМАТЫ_МРФЦ": "Казахстан",
    "МИНСК_МПСЦ": "Беларусь",
}


def parse_excel(file: str = "./excel/report.xlsx") -> list[dict]:
    dataframe = openpyxl.load_workbook(file)
    df = dataframe.active

    cols = []
    for row in range(4, df.max_row):
        data = [col[row].value for col in df.iter_cols(1, df.max_column)]
        values = {
            'sku': data[0],
            'store': data[1],
            'vendor_code': data[2],
            'name': data[3],
            'count_in_transit': data[4],
            'count': data[5],
            'reserve': data[6],
            'idc': data[7],
            'cluster': CLUSTERS[data[1].upper().replace("_НОВЫЙ", "")],
        }
        cols.append(values)
    return cols
